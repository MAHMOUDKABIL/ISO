# -*- coding: utf-8 -*-
"""
أداة اختيارية للتحقق من قالب شيت الفواتير قبل رفعه إلى منصة «مِرْصاد الفواتير».
الاستخدام:  python verify_columns.py مسار_الملف.xlsx
المتطلبات:  pip install openpyxl

إعداد الباحث: محمود الباز فوزي قابيل — لأغراض البحث العلمي فقط
هاتف التواصل: 01067777481
"""
import sys, unicodedata

REQUIRED = [
    "مسلسل", "تفاصيل", "نوع المستند", "نسخة المستند", "الحالة",
    "تاريخ الإصدار", "تاريخ التقديم", "عملة الفاتورة", "قيمة الفاتورة",
    "ضريبة القيمة المضافة", "إجمالى الفاتورة", "الرقم الداخلى", "الرقم الإلكترونى",
]

def normalize(s):
    s = unicodedata.normalize("NFC", str(s or ""))
    for a, b in [("أ","ا"),("إ","ا"),("آ","ا"),("ى","ي"),("ة","ه")]:
        s = s.replace(a, b)
    return "".join(ch for ch in s if not ch.isspace())

def main(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("ثبّت openpyxl أولاً:  pip install openpyxl")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [normalize(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    found, missing = [], []
    for req in REQUIRED:
        if normalize(req) in headers:
            found.append(req)
        else:
            missing.append(req)

    print(f"الملف        : {path}")
    print(f"الورقة       : {ws.title} — {ws.max_row - 1} صف بيانات")
    print(f"أعمدة مطابقة : {len(found)}/{len(REQUIRED)}")
    for h in found:   print("  ✔", h)
    for h in missing: print("  ✘ ناقص:", h)
    sys.exit(0 if not missing else 1)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit("الاستخدام: python verify_columns.py <ملف.xlsx>")
    main(sys.argv[1])